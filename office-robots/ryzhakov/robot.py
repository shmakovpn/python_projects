#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
REQUIREMENTS
python3.7
pip install beautifulsoup4
pip install openpyxl

Задание:

 Открыть URL
 http://10.248.33.133:8888/uralsib/UralSibOldWeb.aspx?polygon=4

 Нажать кнопку
  <input
   name="ButtOtpr"
   type="button"
   id="ButtOtpr"
   value=" Отправление "
   style="width:100%;height:100%;"
   onclick="OnButtClick5()"
   >

 Перейти на открывшуюся вкладку:
  http://10.248.33.133:8080/Otpr_Prib_ob_LD/WF_Otpr_Prib_obh.aspx?&pr_ot=2&poligon=4

 Выбрать НАПРАВЛЕНИЕ: Мариинск-Юрты
  <select
   name="Sel_napr"
   id="Sel_napr"
   style="POSITION: relative; BACKGROUND-COLOR: white; WIDTH: 80%; FONT-SIZE: 10pt"
   tabindex="0"
   >
	<option selected="selected" value="1" pr_sokr_opzd="1">Инская-Мариинск</option>
	<option value="2" pr_sokr_opzd="1">Мариинск-Юрты</option>
	<option value="3" pr_sokr_opzd="1">Юрты-Тайшет</option>
  </select>

 Выбрать ПЕРИОД: тек. сут.
  <select
   name="Sel_per"
   id="Sel_per"
   style="POSITION: relative; BACKGROUND-COLOR: white; WIDTH: 100%; FONT-SIZE: 10pt"
   tabindex="0"
   onchange="go_per();"
   >
	<option selected="selected" value="0">тек. сут.</option>
	<option value="1">отч. сут.</option>
	<option value="2">архив</option>
  </select>

 Нажать кнопку ПУСК (работает только в InternetExplorer)
  <input
   style="BACKGROUND-COLOR: beige; FONT-FAMILY: Arial; HEIGHT: 24px; COLOR: blue; FONT-WEIGHT: bold"
   id="pusc"
   onclick="go_pusc();"
   value="ПУСК"
   type="button"
   >

 Произойдет переход по URL (запрос типа GET)
  http://10.248.33.133:8080/Otpr_Prib_ob_LD/WF_Otpr_Prib_obh.aspx?&Vid_pr=0&pr_ot=2&Period=1&Smena=3&Tip_p=0&kat_p=0&poligon=4&ZZ=0&Dot=0&opozd1=-10&opozd2=5&pr_napr=1
  Заголовки:
   Accept: text/html, application/xhtml+xml, image/jxr, */*
   Accept-Encoding: gzip, deflate
   Accept-Language: en-US, en; q=0.7, ru; q=0.3
   Connection: Keep-Alive
   Cookie: ASP.NET_SessionId=3npe5f55n3zjbw45huj3veqm
   Host: 10.248.33.133:8080
   Referer: http://10.248.33.133:8080/Otpr_Prib_ob_LD/WF_Otpr_Prib_obh.aspx?&Vid_pr=0&pr_ot=2&Period=1&Smena=3&Tip_p=0&kat_p=0&poligon=4&ZZ=0&Dot=0&opozd1=-10&opozd2=5&pr_napr=0
   User-Agent: Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko
  Примечание (залоловки можно игнорировать)
  GET-параметры
   Dot: 0
   kat_p: 0
   opozd1: -10
   opozd2: 5
   Period: 1
   poligon: 4
   pr_napr: 1
   pr_ot: 2
   Smena: 3
   Tip_p: 0
   Vid_pr: 0
   ZZ: 0

 Изменеия в GET-параметры вносить не надо, т.к. по-умолчанию будет выдан требуемый отчет

 Парсинг:
  Будут загружены данные за сутки по данному направлению, необходимо выбрать:
  Колонка 6, итого нечет, итого чет
  Колонка 4,5,7,9 из строки ИТОГО
"""
__author__ = "shmakovpn <shmakovpn@yandex.ru>"
__date__ = "2020-01-28"
import os  # общие функции по работе с операционной системой
from datetime import datetime  # работа с датой и временем
import urllib.parse  # выполнение http запросов
import urllib.request  # выполнение http запросов
from urllib.error import URLError  # обработка ошибок выполнения http запросов
import smtplib  # отправка писем по электронной почте
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from bs4 import BeautifulSoup  # анализатор html
from openpyxl import load_workbook  # работа с файлами excel

CURRENT_DIR = os.path.dirname(__file__)  # папка в которой находится данный скрипт

XLSX_TEMPLATE = 'template.xlsx'  # книга эксель
XLSX_MAX_ROW = 65535  # максимально допустимый номер строки на Листе в Книге Эксель

MAIL_LIST = ('shmakovpn@krw.rzd', 'ar@krw.rzd', )  # список рассылки
MAIL_SERVER = 'uc.krw.rzd'  # адрес почтового сервера
MAIL_PORT = 25  # порт почтового сервера
MAIL_FROM = MAIL_LIST[0]  # от кого сообщение
MAIL_LOCAL_HOSTNAME = 'ivc-ptk-shmak.krw.oao.rzd'  # dns имя компа с которого отправляется почта

NOW = f"{datetime.now():%Y-%m-%d %H:%M:%S}"

# загрузим и разберем web-страницу
url = "http://10.248.33.133:8080/Otpr_Prib_ob_LD/WF_Otpr_Prib_obh.aspx"  # адрес web-страницы с которой берутся данные
values = {
    'Dot': 0,
    'kat_p': 0,
    'opozd1': -10,
    'opozd2': 5,
    'Period': 1,
    'poligon': 4,
    'pr_napr': 1,
    'pr_ot': 2,
    'Smena': 3,
    'Tip_p': 0,
    'Vid_pr': 0,
    'ZZ': 0,
}
data = urllib.parse.urlencode(values)
full_url = f"{url}?{data}"
try:
    with urllib.request.urlopen(full_url) as response:
        the_page = response.read()
        html = the_page.decode()
        soup = BeautifulSoup(html, 'html.parser')
        # Колонка 6, итого нечет
        total_odd = soup.find(id='mainidofrow6idofcell3').getText().replace(',', '.')
        # Колонка 6, итого чет
        total_even = soup.find(id='mainidofrow12idofcell3').getText().replace(',', '.')
        # Колонка 4, ИТОГО
        total_4 = soup.find(id='mainidofrow18idofcell1').getText().replace(',', '.')
        # Колонка 5, ИТОГО
        total_5 = soup.find(id='mainidofrow18idofcell2').getText().replace(',', '.')
        # Колонка 7, ИТОГО
        total_7 = soup.find(id='mainidofrow18idofcell4').getText().replace(',', '.')
        # Колонка 9, ИТОГО
        total_9 = soup.find(id='mainidofrow18idofcell6').getText().replace(',', '.')
except URLError as e:
    print(f"URLErorr e={e}")
    raise URLError(e)
    # todo: продумать логирование ошибки, мониторинг и реакцию на нее (если страницу не удалось загрузить)

# запись данных в excel
try:
    wb = load_workbook(os.path.join(CURRENT_DIR, XLSX_TEMPLATE))
    sheet = wb['trains']
    row = 1
    # находим первую незаполненную строку
    while row < XLSX_MAX_ROW:
        value = sheet.cell(row=row, column=1).value
        if not value:
            break
        row += 1
    if row < XLSX_MAX_ROW:
        sheet.cell(row=row, column=1).value = NOW
        sheet.cell(row=row, column=2).value = total_odd
        sheet.cell(row=row, column=3).value = total_even
        sheet.cell(row=row, column=4).value = total_4
        sheet.cell(row=row, column=5).value = total_5
        sheet.cell(row=row, column=6).value = total_7
        sheet.cell(row=row, column=7).value = total_9
        wb.save(os.path.join(CURRENT_DIR, XLSX_TEMPLATE))  # сохраним изменения в тот же самый файл
        wb.close()
    else:
        # todo предусмотреть возможность превышения максимального номера строки в эксель файле
        raise ValueError('Превышен максимально возможный номер строки эксель файла')
except Exception as e:
    raise Exception(e)  # todo продумать возможные исключения при открытии эксель файла

# отправим полученный файл по электронной почте
msg_content = f"""Роботы идут
datetime            | итого нечетные | итого четные | ИТОГО_4 | ИТОГО_5 | ИТОГО 7 | ИТОГО 9 |
{NOW} | {total_odd} | {total_even} | {total_4} | {total_5} | {total_7} | {total_9} |
"""
msg = MIMEMultipart()
msg['Subject'] = f'Роботы идут {NOW}'
msg['From'] = MAIL_FROM
msg['To'] = ', '.join(MAIL_LIST)
msg['Date'] = formatdate(localtime=True)
msg.attach(MIMEText(msg_content))
with open(os.path.join(CURRENT_DIR, XLSX_TEMPLATE), 'rb') as f:
    attachment = MIMEBase('application', 'octet-stream')
    attachment.set_payload(f.read())
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename="{XLSX_TEMPLATE}"')
    msg.attach(attachment)
with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, local_hostname=MAIL_LOCAL_HOSTNAME) as s:
    s.send_message(msg=msg)

print(f"__END__")



